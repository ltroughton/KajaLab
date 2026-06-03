"""
seahorse_prepare.py
====================
Input: single Excel file with sheets:
  - "Hoescht"  : 8-row x 12-col plate map (rows 1-8), blank rows, then 8x12 fluorescence grid
  - "OCR"      : col A = Time (minutes), cols B+ = well data, row 1 = merged condition headers
  - "ECAR"     : same structure as OCR

Well layout (96-well plate):
  - Plate rows A, H = Background (excluded from OCR/ECAR, used for background subtraction)
  - Inner wells B-G (6 rows) x cols 1-12 = 72 data wells
  - OCR/ECAR columns map positionally: col 1 wells B-G, then col 2 wells B-G, ... col 12 wells B-G

Hoechst processing:
  - Background wells (rows A, H) fluorescence averaged -> subtracted from all inner wells
  - Inner well fluorescence / reference well (B2 by default, = first data well) = scaling factor

Output sheets appended to a new workbook:
  1. Hoescht              - original unchanged
  2. Hoescht_factors      - well ID, raw fluorescence, bg-subtracted, scaling factor
  3. OCR_raw              - original + Cycle column
  4. ECAR_raw             - original + Cycle column  
  5. OCR_adjusted         - Hoechst-normalised OCR
  6. ECAR_adjusted        - Hoechst-normalised ECAR
  7. OCR_corrected        - OCR_adjusted + per-column green/red conditional formatting
  8. ECAR_corrected       - ECAR_adjusted + per-column green/red conditional formatting

Usage:
  python seahorse_prepare.py input.xlsx [output.xlsx]
"""

import sys
import re
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

PLATE_ROWS = list("ABCDEFGH")   # A=0, H=7
PLATE_COLS = list(range(1, 13)) # 1-12
BG_rows    = {"A", "H"}
INNER_ROWS = [r for r in PLATE_ROWS if r not in BG_rows]  # B-G

# ── Styling ──────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", start_color="2F4F8F", end_color="2F4F8F")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT   = Font(name="Arial", size=10)
CENTER      = Alignment(horizontal="center")
THIN        = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"),  bottom=Side(style="thin"))

def style_sheet(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.font      = HEADER_FONT if cell.row == 1 else BODY_FONT
            cell.fill      = HEADER_FILL if cell.row == 1 else PatternFill()
            cell.alignment = CENTER
            cell.border    = THIN
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 30)
    ws.freeze_panes = ws.cell(row=2, column=1)

def write_df(ws, df):
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    style_sheet(ws)

# ── Hoechst parsing ───────────────────────────────────────────────────────────
def parse_hoechst(wb_path, ref_well="B2"):
    ws = load_workbook(wb_path, data_only=True)["Hoescht"]

    # Find where the numeric fluorescence grid starts (first row where A-col is numeric)
    fluor_start_row = None
    for row in ws.iter_rows():
        if row[0].value is not None and isinstance(row[0].value, (int, float)):
            fluor_start_row = row[0].row
            break
    assert fluor_start_row is not None, "Could not find fluorescence data in Hoescht sheet"

    # Read 8x12 fluorescence grid
    fluor = {}
    for r_idx, plate_row in enumerate(PLATE_ROWS):
        for c_idx, plate_col in enumerate(PLATE_COLS):
            cell_val = ws.cell(row=fluor_start_row + r_idx, column=c_idx + 1).value
            fluor[f"{plate_row}{plate_col}"] = float(cell_val) if cell_val is not None else np.nan

    # Background = mean of all A-row and H-row wells
    bg_vals = [v for k, v in fluor.items() if k[0] in BG_rows and not np.isnan(v)]
    bg_mean = np.mean(bg_vals)
    print(f"  Background mean fluorescence: {bg_mean:.1f}")

    # Inner wells: background-subtracted fluorescence
    # Well order matches OCR/ECAR column order: col 1 B-G, col 2 B-G, ... col 12 B-G
    records = []
    well_order = []
    for plate_col in PLATE_COLS:
        for plate_row in INNER_ROWS:
            well = f"{plate_row}{plate_col}"
            raw  = fluor[well]
            corr = raw - bg_mean
            well_order.append(well)
            records.append({"Well": well, "Raw_fluorescence": raw,
                            "BG_subtracted": corr})

    df = pd.DataFrame(records)

    # Scaling factors relative to reference well
    ref_corr = df.loc[df["Well"] == ref_well, "BG_subtracted"].values[0]
    assert ref_corr > 0, f"Reference well {ref_well} has non-positive BG-subtracted fluorescence"
    df["Scaling_factor"] = df["BG_subtracted"] / ref_corr

    print(f"  Reference well: {ref_well}, BG-subtracted fluorescence: {ref_corr:.1f}")

    factors = dict(zip(df["Well"], df["Scaling_factor"]))
    return df, factors, well_order

# ── OCR / ECAR parsing ────────────────────────────────────────────────────────
def parse_ocr_ecar(wb_path, sheet_name, well_order):
    ws = load_workbook(wb_path, data_only=True)[sheet_name]

    # Build column names from merged headers in row 1
    # merged_cells gives us the span of each condition label
    merge_map = {}  # col_index -> condition name
    for mc in ws.merged_cells.ranges:
        if mc.min_row == 1:
            label = ws.cell(row=1, column=mc.min_col).value
            for c in range(mc.min_col, mc.max_col + 1):
                merge_map[c] = label

    # Count replicates per condition (in column order)
    col_names = ["Time (minutes)"]  # col A
    rep_count = {}
    for c_idx in range(2, ws.max_column + 1):  # B onwards
        cond = merge_map.get(c_idx, f"Unknown_col{c_idx}")
        cond = str(cond).strip()
        rep_count.setdefault(cond, 0)
        rep_count[cond] += 1
        col_names.append(f"{cond}_rep{rep_count[cond]}")

    # Read data rows (row 2 onwards)
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        data.append(row[:len(col_names)])

    df = pd.DataFrame(data, columns=col_names)
    df = df.dropna(how="all").reset_index(drop=True)

    # Insert Cycle column before Time
    df.insert(0, "Cycle", range(1, len(df) + 1))

    return df

# ── Normalisation ─────────────────────────────────────────────────────────────
def apply_normalisation(df, factors, well_order):
    df = df.copy()
    rep_cols = [c for c in df.columns if c not in ("Cycle", "Time (minutes)")]

    if len(rep_cols) != len(well_order):
        print(f"  ⚠  {len(rep_cols)} data columns vs {len(well_order)} wells — "
              f"normalising up to min({len(rep_cols)}, {len(well_order)})")

    for i, col in enumerate(rep_cols):
        if i >= len(well_order):
            break
        well   = well_order[i]
        factor = factors[well]
        df[col] = df[col] / factor

    return df

# ── Conditional formatting ────────────────────────────────────────────────────
def add_cf(ws, df):
    skip = {"Cycle", "Time (minutes)"}
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for c_idx, name in enumerate(headers, start=1):
        if name in skip or name is None:
            continue
        col_letter = get_column_letter(c_idx)
        rng = f"{col_letter}2:{col_letter}{ws.max_row}"
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min",        start_color="F8696B",
            mid_type="percentile",   mid_value=50, mid_color="FFFFFF",
            end_type="max",          end_color="63BE7B",
        ))

# ── Main ──────────────────────────────────────────────────────────────────────
def prepare(input_path, output_path=None, ref_well="B2"):
    input_path  = Path(input_path)
    output_path = Path(output_path) if output_path else \
                  input_path.parent / (input_path.stem + "_prepared.xlsx")

    print(f"Input : {input_path}")
    print(f"Output: {output_path}\n")

    print("Parsing Hoechst...")
    factors_df, factors, well_order = parse_hoechst(input_path, ref_well)

    has_ecar = "ECAR" in load_workbook(input_path, read_only=True).sheetnames

    print("Parsing OCR...")
    ocr_raw  = parse_ocr_ecar(input_path, "OCR", well_order)

    if has_ecar:
        print("Parsing ECAR...")
        ecar_raw = parse_ocr_ecar(input_path, "ECAR", well_order)

    print("Normalising...")
    ocr_adj  = apply_normalisation(ocr_raw,  factors, well_order)
    if has_ecar:
        ecar_adj = apply_normalisation(ecar_raw, factors, well_order)

    # Build output workbook
    wb = Workbook()
    wb.remove(wb.active)

    def add_sheet(name, df, cf=False):
        ws = wb.create_sheet(title=name)
        write_df(ws, df)
        if cf:
            add_cf(ws, df)
        print(f"  ✓ {name}")

    print("\nWriting sheets...")
    # Copy original Hoescht sheet as-is
    src_wb = load_workbook(input_path, data_only=True)
    src_ws = src_wb["Hoescht"]
    ws_h = wb.create_sheet("Hoescht")
    for row in src_ws.iter_rows(values_only=True):
        ws_h.append([v for v in row])
    print("  ✓ Hoescht (original)")

    add_sheet("Hoescht_factors", factors_df)
    add_sheet("OCR_raw",         ocr_raw)
    add_sheet("OCR_adjusted",    ocr_adj)
    add_sheet("OCR_corrected",   ocr_adj.copy(), cf=True)

    if has_ecar:
        add_sheet("ECAR_raw",       ecar_raw)
        add_sheet("ECAR_adjusted",  ecar_adj)
        add_sheet("ECAR_corrected", ecar_adj.copy(), cf=True)

    wb.save(output_path)
    print(f"\n✓ Saved: {output_path}")
    print("\nNext steps:")
    print("  1. Open OCR_corrected / ECAR_corrected")
    print("  2. Clear values in any outlier wells (failed injections show as red outliers)")
    print("  3. Run: python seahorse_analyze.py", output_path)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python seahorse_prepare.py input.xlsx [output.xlsx] [ref_well]")
        sys.exit(1)
    prepare(
        sys.argv[1],
        sys.argv[2] if len(sys.argv) > 2 else None,
        sys.argv[3] if len(sys.argv) > 3 else "B2",
    )
